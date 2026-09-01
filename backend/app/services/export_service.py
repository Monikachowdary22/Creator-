from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from openpyxl import Workbook
from openpyxl.styles import Font


# ==========================================
# PDF Export
# ==========================================

def generate_pdf_report(report_data: dict):

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "CreatorIQ Analytics Report",
            styles["Title"]
        )
    )

    elements.append(Spacer(1, 15))

    creator_id = report_data.get(
        "creator_id",
        "N/A"
    )

    elements.append(
        Paragraph(
            f"Creator ID: {creator_id}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 15))

    # ==========================================
    # Content Performance
    # ==========================================

    content_report = report_data.get(
        "content_performance",
        {}
    )

    elements.append(
        Paragraph(
            "Content Performance",
            styles["Heading2"]
        )
    )

    kpi_data = [
        ["KPI", "Value"],
        [
            "Total Content",
            str(content_report.get("total_content", 0))
        ],
        [
            "Total Views",
            str(content_report.get("total_views", 0))
        ],
        [
            "Total Likes",
            str(content_report.get("total_likes", 0))
        ],
        [
            "Total Comments",
            str(content_report.get("total_comments", 0))
        ],
        [
            "Total Shares",
            str(content_report.get("total_shares", 0))
        ],
        [
            "Total Reach",
            str(content_report.get("total_reach", 0))
        ]
    ]

    table = Table(kpi_data)

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("PADDING", (0, 0), (-1, -1), 6)
        ])
    )

    elements.append(table)

    elements.append(Spacer(1, 15))

    # ==========================================
    # Content Table
    # ==========================================

    contents = content_report.get(
        "content",
        []
    )

    if contents:

        elements.append(
            Paragraph(
                "Content Details",
                styles["Heading2"]
            )
        )

        content_table = [
            [
                "ID",
                "Title",
                "Platform",
                "Views",
                "Likes",
                "Comments",
                "Reach"
            ]
        ]

        for content in contents:

            content_table.append([
                str(content.get("id", "")),
                str(content.get("title", ""))[:30],
                str(content.get("platform", "")),
                str(content.get("views", 0)),
                str(content.get("likes", 0)),
                str(content.get("comments", 0)),
                str(content.get("reach", 0))
            ])

        table = Table(
            content_table,
            repeatRows=1
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.grey
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.black
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    7
                ),
                (
                    "PADDING",
                    (0, 0),
                    (-1, -1),
                    4
                )
            ])
        )

        elements.append(table)

        elements.append(Spacer(1, 15))

    # ==========================================
    # Revenue
    # ==========================================

    revenue_report = report_data.get(
        "revenue_analytics",
        {}
    )

    elements.append(
        Paragraph(
            "Revenue Analytics",
            styles["Heading2"]
        )
    )

    revenue_table = Table([
        ["Metric", "Value"],
        [
            "Total Revenue",
            str(
                revenue_report.get(
                    "total_revenue",
                    0
                )
            )
        ],
        [
            "Revenue Records",
            str(
                revenue_report.get(
                    "total_records",
                    0
                )
            )
        ]
    ])

    revenue_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("PADDING", (0, 0), (-1, -1), 6)
        ])
    )

    elements.append(revenue_table)

    elements.append(Spacer(1, 15))

    # ==========================================
    # Platform Comparison
    # ==========================================

    platforms = report_data.get(
        "platform_comparison",
        []
    )

    if platforms:

        elements.append(
            Paragraph(
                "Platform Comparison",
                styles["Heading2"]
            )
        )

        platform_table = [
            [
                "Platform",
                "Content",
                "Views",
                "Reach",
                "Engagement Rate"
            ]
        ]

        for platform in platforms:

            platform_table.append([
                str(platform.get("platform", "")),
                str(platform.get("content_count", 0)),
                str(platform.get("total_views", 0)),
                str(platform.get("total_reach", 0)),
                str(platform.get("engagement_rate", 0))
                + "%"
            ])

        table = Table(
            platform_table,
            repeatRows=1
        )

        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("PADDING", (0, 0), (-1, -1), 5)
            ])
        )

        elements.append(table)

    # ==========================================
    # Sponsorships
    # ==========================================

    sponsorship_report = report_data.get(
        "sponsorships",
        {}
    )

    sponsorships_data = sponsorship_report.get(
        "data",
        []
    )

    if sponsorships_data:
        elements.append(Spacer(1, 15))

        elements.append(
            Paragraph(
                "Sponsorship Campaigns",
                styles["Heading2"]
            )
        )

        sponsorship_table = [
            [
                "Brand",
                "Campaign",
                "Contract Value",
                "Status",
                "Payment"
            ]
        ]

        for sp in sponsorships_data:
            sponsorship_table.append([
                str(sp.get("brand_name", "")),
                str(sp.get("campaign", "")),
                str(sp.get("contract_value", 0)),
                str(sp.get("status", "")),
                str(sp.get("payment_status", ""))
            ])

        sp_table = Table(
            sponsorship_table,
            repeatRows=1
        )

        sp_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("PADDING", (0, 0), (-1, -1), 5)
            ])
        )

        elements.append(sp_table)

    document.build(elements)

    buffer.seek(0)

    return buffer


# ==========================================
# Excel Export
# ==========================================

def generate_excel_report(report_data: dict):

    workbook = Workbook()

    # ==========================================
    # Content Sheet
    # ==========================================

    sheet = workbook.active
    sheet.title = "Content Performance"

    content_report = report_data.get(
        "content_performance",
        {}
    )

    sheet.append([
        "Creator ID",
        report_data.get("creator_id", "")
    ])

    sheet.append([])

    sheet.append([
        "Total Content",
        content_report.get("total_content", 0)
    ])

    sheet.append([
        "Total Views",
        content_report.get("total_views", 0)
    ])

    sheet.append([
        "Total Likes",
        content_report.get("total_likes", 0)
    ])

    sheet.append([
        "Total Comments",
        content_report.get("total_comments", 0)
    ])

    sheet.append([
        "Total Shares",
        content_report.get("total_shares", 0)
    ])

    sheet.append([
        "Total Reach",
        content_report.get("total_reach", 0)
    ])

    sheet.append([])

    sheet.append([
        "ID",
        "Title",
        "Platform",
        "Views",
        "Likes",
        "Comments",
        "Shares",
        "Saves",
        "Reach",
        "Published Date"
    ])

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for cell in sheet[10]:
        cell.font = Font(bold=True)

    for content in content_report.get(
        "content",
        []
    ):

        sheet.append([
            content.get("id", ""),
            content.get("title", ""),
            content.get("platform", ""),
            content.get("views", 0),
            content.get("likes", 0),
            content.get("comments", 0),
            content.get("shares", 0),
            content.get("saves", 0),
            content.get("reach", 0),
            str(content.get("published_date", ""))
        ])

    # ==========================================
    # Revenue Sheet
    # ==========================================

    revenue_sheet = workbook.create_sheet(
        "Revenue Analytics"
    )

    revenue_report = report_data.get(
        "revenue_analytics",
        {}
    )

    revenue_sheet.append([
        "Total Revenue",
        revenue_report.get(
            "total_revenue",
            0
        )
    ])

    revenue_sheet.append([
        "Total Records",
        revenue_report.get(
            "total_records",
            0
        )
    ])

    revenue_sheet.append([])

    revenue_sheet.append([
        "ID",
        "Amount",
        "Date",
        "Source",
        "Description"
    ])

    for cell in revenue_sheet[4]:
        cell.font = Font(bold=True)

    for revenue in revenue_report.get(
        "data",
        []
    ):

        revenue_sheet.append([
            revenue.get("id", ""),
            revenue.get("amount", 0),
            str(revenue.get("revenue_date", "")),
            revenue.get("source", ""),
            revenue.get("description", "")
        ])

    # ==========================================
    # Platform Sheet
    # ==========================================

    platform_sheet = workbook.create_sheet(
        "Platform Comparison"
    )

    platform_sheet.append([
        "Platform",
        "Content Count",
        "Total Views",
        "Total Likes",
        "Total Comments",
        "Total Shares",
        "Total Reach",
        "Total Engagement",
        "Engagement Rate"
    ])

    for cell in platform_sheet[1]:
        cell.font = Font(bold=True)

    for platform in report_data.get(
        "platform_comparison",
        []
    ):

        platform_sheet.append([
            platform.get("platform", ""),
            platform.get("content_count", 0),
            platform.get("total_views", 0),
            platform.get("total_likes", 0),
            platform.get("total_comments", 0),
            platform.get("total_shares", 0),
            platform.get("total_reach", 0),
            platform.get("total_engagement", 0),
            platform.get("engagement_rate", 0)
        ])

    # ==========================================
    # Growth Sheet
    # ==========================================

    growth_sheet = workbook.create_sheet(
        "Growth Trends"
    )

    growth_report = report_data.get(
        "growth_trends",
        {}
    )

    growth_data = growth_report.get(
        "data",
        []
    )

    if growth_data:

        headers = list(
            growth_data[0].keys()
        )

        growth_sheet.append(headers)

        for cell in growth_sheet[1]:
            cell.font = Font(bold=True)

        for record in growth_data:

            growth_sheet.append([
                record.get(header, "")
                for header in headers
            ])

    else:

        growth_sheet.append([
            "Growth Trends"
        ])

        growth_sheet.append([
            "No growth records available"
        ])

    # ==========================================
    # Audience Sheet
    # ==========================================

    audience_sheet = workbook.create_sheet(
        "Audience Analytics"
    )

    audience_report = report_data.get(
        "audience_analytics",
        {}
    )

    audience_data = audience_report.get(
        "data",
        []
    )

    if audience_data:

        headers = list(
            audience_data[0].keys()
        )

        audience_sheet.append(headers)

        for cell in audience_sheet[1]:
            cell.font = Font(bold=True)

        for record in audience_data:

            audience_sheet.append([
                record.get(header, "")
                for header in headers
            ])

    else:

        audience_sheet.append([
            "Audience Analytics"
        ])

        audience_sheet.append([
            "No audience records available"
        ])

    # ==========================================
    # Sponsorships Sheet
    # ==========================================

    sponsorship_sheet = workbook.create_sheet(
        "Sponsorships"
    )

    sponsorship_report = report_data.get(
        "sponsorships",
        {}
    )

    sponsorship_data = sponsorship_report.get(
        "data",
        []
    )

    if sponsorship_data:

        headers = list(
            sponsorship_data[0].keys()
        )

        sponsorship_sheet.append(headers)

        for cell in sponsorship_sheet[1]:
            cell.font = Font(bold=True)

        for record in sponsorship_data:

            sponsorship_sheet.append([
                record.get(header, "")
                for header in headers
            ])

    else:

        sponsorship_sheet.append([
            "Sponsorships"
        ])

        sponsorship_sheet.append([
            "No sponsorship records available"
        ])

    # ==========================================
    # Adjust Column Widths
    # ==========================================

    for worksheet in workbook.worksheets:

        for column in worksheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40
            )

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer